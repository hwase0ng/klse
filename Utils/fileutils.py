'''
Created on Dec 20, 2016

@author: t.roy
'''
from datetime import datetime
import mmap
import openpyxl
import subprocess
import xlrd

def mapcount(filename):
    f = open(filename, "r+")
    buf = mmap.mmap(f.fileno(), 0)
    lines = 0
    readline = buf.readline
    while readline():
        lines += 1
    return lines

def wc_line_count(filename):
    out = subprocess.Popen(['wc', '-l', filename],
                         stdout=subprocess.PIPE,
                         stderr=subprocess.STDOUT
                         ).communicate()[0]
    return int(out.partition(b' ')[0])

def tail(fl, n=1, bs=1024):
    try:
        f = open(fl)
        f.seek(-1,2)
        l = 1-f.read(1).count('\n') # If file doesn't end in \n, count it anyway.
        B = f.tell()
        while n >= l and B > 0:
                block = min(bs, B)
                B -= block
                f.seek(B, 0)
                l += f.read(block).count('\n')
        f.seek(B, 0)
        l = min(l,n) # discard first (incomplete) line if l > n
        lines = f.readlines()[-l:]
        f.close()
        return lines
    except:
        return [""]
    
def xls_to_xlsx(*args, **kw):
    """
    open and convert an XLS file to openpyxl.workbook.Workbook
    ----------
    @param args: args for xlrd.open_workbook
    @param kw: kwargs for xlrd.open_workbook
    @return: openpyxl.workbook.Workbook
    """
    book_xls = xlrd.open_workbook(*args, formatting_info=True, ragged_rows=True, **kw)
#   book_xlsx = openpyxl.workbook.Workbook()
    book_xlsx = openpyxl.Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])
        for crange in sheet_xls.merged_cells:
            rlo, rhi, clo, chi = crange
            sheet_xlsx.merge_cells(start_row=rlo + 1, end_row=rhi,
            start_column=clo + 1, end_column=chi,)

        def _get_xlrd_cell_value(cell):
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                datetime_tup = xlrd.xldate_as_tuple(value,0)    
                if datetime_tup[0:3] == (0, 0, 0):   # time format without date
                    value = datetime.time(*datetime_tup[3:])
                else:
                    value = datetime.datetime(*datetime_tup)
            return value

        for row in range(sheet_xls.nrows):
            sheet_xlsx.append((
                _get_xlrd_cell_value(cell)
                for cell in sheet_xls.row_slice(row, end_colx=sheet_xls.row_len(row))
            ))
    return book_xlsx

if __name__ == '__main__':
    pass